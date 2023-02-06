# -*- coding: utf-8 -*-
#Python Version 3.8

import sys,os,random,importlib
import configparser
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,PatternFill,Alignment
from openpyxl import load_workbook
import urllib.request,json


'''
Create by huanghaiyang 2019.11.26

版本:1.5.2 黄海洋 2020-03-23
写入txt文件时如遇非法字符则报错，提示具体哪行哪列

版本:1.6 黄海洋 2020-03-24
认购金额之前采用的是实际金额乘以100,现在按实际的金额来处理

版本:1.7 黄海洋 2020-05-07
金额拆分逻辑由原来的按固定值拆分，修改为按随机值拆分

版本:2.0 黄海洋 2020-05-25
增加快付通代付txt文件格式的转换

版本:2.1 黄海洋 2020-06-08
增加银联单笔限额500万拆分

版本:2.2 黄海洋 2020-06-30
由原来python 2.7改为python 3.8版本进行编程
修改快付通行别代码，个人客户行别代码可为空

版本:3.0 黄海洋 2022-01-17
增加查询手机号、身份证和银行卡的功能
'''


'''
随机拆分一个正整数，
返回list数组
'''
def RandomSplit(Number,Max_Int,Min_Int,Split_amount):
    '''
    Number        #待拆分数字
    Max_Int    #随机数上限
    Min_Int    #随机数下限
    Split_amount  #银联代付单笔上限
    '''
    list = []

    if Number >= Split_amount:                       #小于100万*100不拆分
        while Number > Max_Int*1000000:
            num_r=random.randint(Min_Int,Max_Int)*1000000   #生成随机数值
            list.append(num_r)                      #list追加数值
            Number = Number - num_r                 #减法
        list.append(Number)                         #list追加剩余的数值

    else:
        list.append(Number)                         #参数Number如果小于等于Max_Int，直接返回参数值

    return list


def CopyFile():
    '''
    复制Excel文件
    '''
    sSourceFile="data.xlsx"
    sTargetFile="target.xlsx"
    wb = openpyxl.load_workbook(sSourceFile)
    #copy_sheet1=wb.copy_worksheet(wb.worksheets[0])
    wb.save(sTargetFile)

##    print("=====copy file <<<target.xlsx>>> completed!=====\n")

def DeleteNoneRows():
    '''
    删除EXCEL表中的空行
    '''
    FileName = "target.xlsx"
    wb = load_workbook(FileName)
    ws = wb.worksheets[0]
    rows = ws.max_row #excel行数
    column = ws.max_column #excel列数
    empty_rows = []

    #取空行的rowid
    for i in range(4,rows+1):
        if (ws.cell(i,4).value is None and \
            ws.cell(i,5).value is None and \
            ws.cell(i,10).value is None):
##            print "test none"+ str(i)
            empty_rows.append(i)

    for i in reversed(empty_rows): #倒序删除excel空行
        ws.delete_rows(i)

    wb.save(FileName)


def ModFile(Cmd_no):
    '''
    拆分和修改target.xlsx文件
    '''
    print("\n")
    print("========================================================")
    print("*  注意:data.xlsx文件中客户认购金额请使用**常规**格式  *")
    print("========================================================")
    print("\n")
    FileName = "target.xlsx"
    List_tmp,List_cust,List_AmountMoney = [],[],[]
    List_num = [] #正整数拆分后返回的list
    AmountMoney = 0
    Divisor = 99000000  #除数
    wb = load_workbook(FileName)
    ws = wb.worksheets[0]
    rows = ws.max_row #excel行数
    column = ws.max_column #excel列数

    #初始化拆分参数
    if Cmd_no =='1':        #银联单笔500万限额
        Max_Int=499
        Min_Int=400
        Split_amount=500000000
    elif Cmd_no =='2':       #银联单笔100万限额
        Max_Int=99
        Min_Int=88
        Split_amount=100000000
    else:
        print("初始化拆分参数错误，命令参数错误！！")
        os.system("pause")
        sys.exit()


    #修改认购金额，认购金额乘以100以浮点格式保留2位，再转为整型
    for i in range(rows,3,-1):
        num_tmp = int(float('%.2f' %((ws.cell(i,10).value)*100)))
        ws.cell(i,10).value = num_tmp
##        print ws.cell(i,10).number_format


    #取最大客户号
    for i in range(4,rows+1):
        List_cust.append(ws.cell(i,3).value) #客户号
        List_AmountMoney.append(ws.cell(i,10).value) #金额
    CustometMaxId = int(max(List_cust))
    Count_num =  len(List_cust) #笔数


    #计算总金额
    for i in range(0,len(List_AmountMoney)):
        if List_AmountMoney[i] is not None:
            AmountMoney+=int(List_AmountMoney[i])
   
    #拆分
    for i in range(rows,3,-1):
        if ws.cell(i,10).value >= Split_amount:  #金额大于等于单笔限额拆分
            print("第%s行 认购金额大于单笔限额!准备拆分..." %(i))

            '''
            #拆分逻辑修改为随机拆分，此段相同数值拆分代码作废
            try:
                num = ws.cell(i,10).value / Divisor
##                print num
                num_mod = ws.cell(i,10).value % Divisor
##            print num_mod
            except:
                print u"错误:请确认认购金额是否为数字格式!"
                os.system("pause")
                sys.exit()

            ws.cell(i,10).value = num_mod
            ws.cell(i,10).font = Font(color=colors.RED)
            ws.cell(i,10).fill = PatternFill("solid", fgColor=colors.YELLOW)
            '''

            List_num = RandomSplit(ws.cell(i,10).value,Max_Int,Min_Int,Split_amount) #调用随机拆分正整数函数
            num=len(List_num)    #拆分后数值的个数

            #修改原该行中的金额数值
            ws.cell(i,10).value = List_num[0]
            ws.cell(i,10).font = Font(color=colors.RED)
            ws.cell(i,10).fill = PatternFill("solid", fgColor=colors.YELLOW)


            for j in range(1,column+1):
                if j==10:
                    List_tmp.append(Divisor)      #无实际意义 金额后面重新赋值
                else:
                    List_tmp.append(ws.cell(i,j).value)
##            List_tmp.append("Source:"+ws.cell(i,3).value) #追加源客户号
            CustomId_tmp=str(ws.cell(i,3).value) #暂存客户号

            #客户号2位子号限制金额不可过大
            if num > 99:
                print("错误:第%s行金额过大,单笔100W不可大于%s万元,单笔500W不可大于%s万元!" %(i,88000000*99/100,400000000*99/100))
                os.system("pause")
                sys.exit()

            #excel表追加数据
            for k in range (1,num):
##                CustometMaxId = CustometMaxId + 1
                Count_num = Count_num + 1  #总笔数
                List_tmp[2]=CustomId_tmp + str(num-k).zfill(2)  #客户号追加子号
                List_tmp[9]=List_num[k]        #将拆分后的正整数赋给金额列
                ws.insert_rows(i)
                for m in range(1,column+1):
                    ws.cell(i,m).value = List_tmp[m-1]    #新行插入数据
                ws.cell(i,10).font = Font(color=colors.RED)
                ws.cell(i,10).fill = PatternFill("solid", fgColor=colors.YELLOW)
##                ws.append(List_tmp)

            print("第%s行 格式转换和拆分完成!\n" %(i))

            List_tmp = [] #重置

    #修改总笔数和总金额
##    print Count_num
    ws.cell(2,5).value=Count_num
    ws.cell(2,6).value=AmountMoney

##    ws.auto_filter.add_sort_condition("D4:D40")
    wb.save(FileName)
##    print("=====modify file <<<target.xlsx>>> completed!=====\n")


def ExcelToTxt(Cmd_no):
    '''
    银联格式
    生成txt文件和新excel文件
    删除临时文件target.xlsx
    '''
    ExcelFileName="target.xlsx"
    wb = load_workbook(ExcelFileName)
    ws = wb.worksheets[0]
    rows = ws.max_row #行数
    
##    MerchantNumber = ws['C2'].value    #取商户号
    #初始化商户号
    if Cmd_no =='1':
        MerchantNumber = 808080211308486
    elif Cmd_no =='2':
        MerchantNumber = 808080211308079
    else:
        print("初始化商户号错误，命令参数错误！！")
        os.system("pause")
        sys.exit()
        
    Date = ws['B4'].value   #取商户日期
    BatchNumber = ws['D2'].value    #取批次号
    TxtFileName = str(MerchantNumber).strip() + '_'+ str(Date).strip() + '_'+\
                  str(BatchNumber).strip() + '.txt'


    stdi,stdo,stde=sys.stdin,sys.stdout,sys.stderr
##    reload(sys)  
    importlib.reload(sys)
    sys.stdin,sys.stdout,sys.stderr=stdi,stdo,stde #解决reload后无输出问题
##    sys.setdefaultencoding('GBK') 
    with open(TxtFileName,"w") as txt:
        #文件头
        txt.write(str(MerchantNumber).strip()+"|"+str(ws['D2'].value).strip()+"|"+str(ws['E2'].value).strip()+"|"+str(ws['F2'].value).strip()+"\n")
        #写入数据
        for i in range(4,rows+1):
            for j in range(2,12):
##                print str(ws.cell(i,j).value)
                try:
                    txt.write(str(ws.cell(i,j).value).strip() +"|")
                except Exception as e:
                    print("临时文件target第%s行第%s列数据有误，可能存在非法字符，请检查！" %(i,j))
                    print(e)
                    os.system("pause")
                    sys.exit()
            if i != rows:
                txt.write("\n")
##    txt.flush()
##    txt.close()

    #按新的文件名生成excel
    ExcelFileName = str(MerchantNumber).strip() + '_'+ str(Date).strip() + '_'+\
                      str(BatchNumber).strip() + '.xlsx'
    wb.save(ExcelFileName)
    os.remove("target.xlsx") #删除临时文件
##    print("=====delete file <<<target.xlsx>>> completed!=====\n")
    print("提示信息：新建Txt文件:"+ TxtFileName)
    print("提示信息：新建Excel文件:"+ ExcelFileName)
    print("\n=====  恭喜你:data.xlsx文件转换拆分为【银联_%s】格式完成!  =====\n" %(MerchantNumber))
    print("\n=====  温馨提示:转换后的文件还需人工仔细核对!  =====\n")



def GetBankcode_kmt(bankname):
    '''
    将银行名称修改为快付通的银行代码
    bankconde.ini保存为ANSI编码格式
    '''
    #  实例化configParser对象
    config = configparser.ConfigParser()
    # -read读取ini文件
    #config.read('bankcode.ini',encoding='utf-8')
    config.read('bankcode.ini')
    # -sections得到所有的section，并以列表的形式返回
    #print('sections:' , ' ' , config.sections())
    # -options(section)得到该section的所有option
    #print('options:' ,' ' , config.options('bankcode'))

    # -items（section）得到该section的所有键值对
    #print('items:' ,' ' ,config.items('bankcode'))

    # -get(section,option)得到section中option的值，返回为string类型
    #print('get:' ,' ' , config.get('bankcode', u'工商银行'))
    try:
##        bankcode=config.get('bankcode_kft', u'中国工商银行股份有限公司1')
        bankcode=config.get('bankcode_kft', bankname)
    except:
        bankcode='null'
    return bankcode


def ModifyBankcode_kmt():
    print("\n")
    print("==============================================================")
    print("*  注意:data.xlsx文件中客户认购金额请使用**常规**格式        *")
    print("*  注意:快付通格式,机构客户【行别代码】必填，个人客户可为空  *")
    print("==============================================================")
    print("\n")
    FileName = "target.xlsx"
    List_tmp,List_cust,List_AmountMoney = [],[],[]
    List_num = [] #正整数拆分后返回的list
    AmountMoney = 0
    Divisor = 99000000  #除数
    wb = load_workbook(FileName)
    ws = wb.worksheets[0]
    rows = ws.max_row #excel行数
    column = ws.max_column #excel列数
    error_count = 0    #查询银行代码错误计数
    error_count_jg =0   #查询银行代码错误计数_疑似机构

    #修改认购金额，认购金额乘以100以浮点格式保留2位，再转为整型
    for i in range(rows,3,-1):
        num_tmp = int(float('%.2f' %((ws.cell(i,10).value)*100)))
        ws.cell(i,10).value = num_tmp
##        print ws.cell(i,10).number_format


    #取最大客户号
    for i in range(4,rows+1):
        List_cust.append(ws.cell(i,3).value) #客户号
        List_AmountMoney.append(ws.cell(i,10).value) #金额
    CustometMaxId = int(max(List_cust))
    Count_num =  len(List_cust) #笔数


    #计算总金额
    for i in range(0,len(List_AmountMoney)):
        if List_AmountMoney[i] is not None:
            AmountMoney+=int(List_AmountMoney[i])


    #将银行名称修改为快付通银行代码
    for i in range(rows,3,-1):
        if ws.cell(i,10).value >= 1000000000:  #金额大于等于1000万
            print("第%s行 认购金额大于1000万元,超过单笔代付限额!" %(i))
            os.system("pause")
            sys.exit()

        else:
            bankname=ws.cell(i,6).value  #取开户行名称
            bankcode=GetBankcode_kmt(bankname)        #获取银行代码
            cust_name=ws.cell(i,5).value    #取客户姓名
            if bankcode=='null':
                ws.cell(i,6).value=' '      #未查询到行别代码则赋空值
                if len(cust_name) > 5:      #客户姓名长度大于5的,列为"疑似机构"
                    print("提示信息：第%s行,客户姓名【%s】,行别名称【%s】疑似机构，未查询到银行代码!" %(i,cust_name,bankname))
                    error_count_jg = error_count_jg + 1
                error_count = error_count + 1
            else: 
                ws.cell(i,6).value=bankcode           #赋值银行代码
            ws.cell(i,3).value=i-3                    #客户号改为序号
            '''
            List_num = RandomSplit(ws.cell(i,10).value) #调用随机拆分正整数函数
            num=len(List_num)    #拆分后数值的个数

            #修改原该行中的金额数值
            ws.cell(i,10).value = List_num[0]
            ws.cell(i,10).font = Font(color=colors.RED)
            ws.cell(i,10).fill = PatternFill("solid", fgColor=colors.YELLOW)


            for j in range(1,column+1):
                if j==10:
                    List_tmp.append(Divisor)      #无实际意义 金额后面重新赋值
                else:
                    List_tmp.append(ws.cell(i,j).value)
##            List_tmp.append("Source:"+ws.cell(i,3).value) #追加源客户号
            CustomId_tmp=str(ws.cell(i,3).value) #暂存客户号

            #客户号2位子号限制金额不可过大
            if num > 99:
                print u"错误:第%s行金额过大,不可大于%s万元!" %(i,99000000*99/100)
                os.system("pause")
                sys.exit()

            #excel表追加数据
            for k in range (1,num):
##                CustometMaxId = CustometMaxId + 1
                Count_num = Count_num + 1  #总笔数
                List_tmp[2]=CustomId_tmp + str(num-k).zfill(2)  #客户号追加子号
                List_tmp[9]=List_num[k]        #将拆分后的正整数赋给金额列
                ws.insert_rows(i)
                for m in range(1,column+1):
                    ws.cell(i,m).value = List_tmp[m-1]    #新行插入数据
                ws.cell(i,10).font = Font(color=colors.RED)
                ws.cell(i,10).fill = PatternFill("solid", fgColor=colors.YELLOW)
##                ws.append(List_tmp)

            print u"第%s行 格式转换和拆分完成!\n" %(i)

            List_tmp = [] #重置
            '''
    if error_count > 0:
        print("提示信息：共计【%s】条记录未查询到银行代码!其中【%s】条疑似机构！"  %(error_count,error_count_jg))
        print("提示信息：机构客户【行别代码】必填,个人客户可为空.请查看和修改银行代码配置文件bankcode.ini \n")
##        os.system("pause")
##        sys.exit()      
    
    #修改总笔数和总金额
    ws.cell(2,5).value=Count_num
    ws.cell(2,6).value=AmountMoney

##    ws.auto_filter.add_sort_condition("D4:D40")
    wb.save(FileName)
##    print("=====modify file <<<target.xlsx>>> completed!=====\n")


def ExcelToTxt_kft():
    '''
    生成快付通txt文件和新excel文件
    删除临时文件target.xlsx
    '''
    ExcelFileName="target.xlsx"
    wb = load_workbook(ExcelFileName)
    ws = wb.worksheets[0]
    rows = ws.max_row #行数
    MerchantNumber = '2020051803488720'    #取商户号
    Date = str(ws['B4'].value).strip()   #取商户日期
    BatchNumber = str(ws['D2'].value).strip()    #取批次号
    Business_type = '11'      #业务类型批量代付'11'
    Service_number = 'BTPDF002'  #服务编号
    
    TxtFileName = str(BatchNumber).strip() + '.txt'

    stdi,stdo,stde=sys.stdin,sys.stdout,sys.stderr
##    reload(sys)  #python2
    importlib.reload(sys)
    sys.stdin,sys.stdout,sys.stderr=stdi,stdo,stde #解决reload后无输出问题
##    sys.setdefaultencoding('GBK') #python2
    
    with open(TxtFileName,"w") as txt:
        #文件头_快付通
        txt.write(Business_type + str(MerchantNumber).ljust(20,' ') + str(BatchNumber).ljust(32,' ') + str(Service_number).ljust(8,' ') + \
                  Date +  str(ws['E2'].value).strip().rjust(8,'0') + str(ws['F2'].value).strip().rjust(16,'0') + ''.ljust(32,' ') + "\n")
        #写入数据
        for i in range(4,rows+1):
            try:
                txt.write(
                          str(BatchNumber+str(ws.cell(i,3).value).strip().rjust(6,'0')).ljust(32,' ') + ''.ljust(32,' ') + str(ws.cell(i,10).value).strip().rjust(16,'0') + \
                          #订单编号、协议编号、收款金额
                           '代付'.ljust(64,' ') + str(ws.cell(i,6).value).strip().ljust(7,' ') + '1' + ' '+ str(ws.cell(i,4).value).strip().ljust(32,' ') + \
                          #交易名称、银行行别、客户银行账户类型、银行卡类型、银行账号
                           str(ws.cell(i,5).value).strip().ljust(64,' ') +  \
                          #姓名
                          ''.ljust(32,' ') + ''.ljust(20,' ') + ''.ljust(4,' ') + ''.ljust(3,' ') + \
                           #证件号、手机号、客户信用卡有效期、客户信用卡的cvv2
                           str(ws.cell(i,11).value).strip().ljust(128,' ')
                          #备注
                         )
                #本程序中，ljust和rjust填充包含中英文字符串需要decode('gbk')，不然填充后长度不对(python3无需decode)
                           
            except Exception as e:
                print("临时文件target第%s行数据有误，可能存在非法字符，请检查！" %(i))
                print(e)
                os.system("pause")
                sys.exit()
            if i != rows:
                txt.write("\n")
##    txt.flush()
##    txt.close()

    #按新的文件名生成excel
    ExcelFileName = str(BatchNumber).strip() + '.xlsx'
    wb.save(ExcelFileName)
    os.remove("target.xlsx") #删除临时文件
##    print("=====delete file <<<target.xlsx>>> completed!=====\n")
    print("提示信息：新建Txt文件:"+ TxtFileName)
    print("提示信息：新建Excel文件:"+ ExcelFileName)
    print("\n=====  恭喜你:data.xlsx文件转换为【快付通】格式完成!  =====\n")
    print("\n=====  温馨提示:转换后的文件还需人工仔细核对!  =====\n")


def GetData(para):
    '''
    查询接口
    该接口有数据格式校验，因此本程序不校验数据格式
    '''
    baseurl = 'http://www.zhaotool.com/v1/api/lt/e10adc3949ba59abbe56e057f20f883e/'
    newurl = baseurl + para
    res = {}
    try:
        with urllib.request.urlopen(newurl) as f:
            res = json.loads(f.read().decode('utf-8'))
    except urllib.error.URLError as e:
        print('程序调用接口失败：%s' % e.reason)
    except:
        print('ERROR：程序异常！请检查输入的内容格式是否正确，或者联系管理员！')
    return res


if __name__=="__main__":
    while True:
        print('\n                 欢迎使用文件格式转换程序\n')
        print('===================================================================')
        print('*  1.文件data.xlsx转换为【银联_500万限额_808080211308486】格式    *')
        print('*  2.文件data.xlsx转换为【银联_100万限额_808080211308079】格式    *')
        print('*  3.文件data.xlsx转换为【快付通】格式                            *')
        print('*  4.查询【手机号】归属                                           *')
        print('*  5.查询【身份证号】归属                                         *')
        print('*  6.查询【银行卡号】归属                                         *')
        print('*  0.退出                                                         *')
        print('*                                                                 *')
        print('*                                             Version:3.0         *')
        print('*                                        Create by Huanghaiyang   *')
        print('===================================================================')
        print('\n')
        print('请输入命令编号')

        cmd_no = input('Please input command No:')

        if cmd_no == '1':
            CopyFile()
            DeleteNoneRows()
            ModFile('1')
            ExcelToTxt('1')
            os.system("pause")
            sys.exit()

        elif cmd_no == '2':
            CopyFile()
            DeleteNoneRows()
            ModFile('2')
            ExcelToTxt('2')
            os.system("pause")
            sys.exit()

        elif cmd_no == '3':
            CopyFile()
            DeleteNoneRows()
            ModifyBankcode_kmt()
            ExcelToTxt_kft()
            os.system("pause")
            sys.exit()

        elif cmd_no == '4':
            mobile_no =  input('请输入手机号：')
            res = GetData(mobile_no)    #调用查询接口 测试用例：18610635264
            try:
                if res['code']=='0':
                    print('===================================================================')
                    print('* 手机号：%s'% res['data']['mobile'])
                    print('* 归属地：%s'% res['data']['phoneArea'])
                    print('===================================================================')
                else:
                    print('===================================================================')
                    print('* %s' % res)
                    print('===================================================================')
                os.system("pause")
            except:
                pass
                os.system("pause")
##              sys.exit()

        elif cmd_no == '5':
            id_no =  input('请输入身份证号：')
            res = GetData(id_no)    #调用查询接口 测试用例：310108198609140018
            try:
                if res['code']=='0':
                    print('===================================================================')
                    print('* 生  日：%s'% res['data']['birthday'])
                    print('* 归属地：%s'% res['data']['area'])
                    print('* 星  座：%s'% res['data']['constellation'])
                    print('* 属  相：%s'% res['data']['zodiac'])
                    print('* 身份证：%s'% res['data']['idcard'])
                    print('* 性  别：%s'% res['data']['sex'])
                    print('* 农历年：%s'% res['data']['cnEra'])
                    print('===================================================================')
                else:
                    print('===================================================================')
                    print('* %s' % res)
                    print('===================================================================')
                os.system("pause")
            except:
                pass
                os.system("pause")
##              sys.exit()

        elif cmd_no == '6':
            bank_no =  input('请输入银行卡号：')
            res = GetData(bank_no)    #调用查询接口 测试用例：6210300009958544
            try:
                if res['code']=='0':
                    if res['data']:
                        print('===================================================================')
                        print('* 银行名称：%s'% res['data']['bankName'])
                        print('* 卡 类 型：%s'% res['data']['cardType'])
                        print('* 银行网址：%s'% res['data']['site'])
                        print('* 客服电话：%s'% res['data']['bankMobile'])
                        print('* 银行卡号：%s'% res['data']['bankCard'])
                        print('===================================================================')
                    else:
                        print('===================================================================')
                        print('* 没有查询到数据！')
                        print('===================================================================')
                else:
                    print('===================================================================')
                    print('* %s' % res)
                    print('===================================================================')
                os.system("pause")
            except:
                pass
                os.system("pause")
##              sys.exit()               
            
        elif cmd_no == '0':
            break

        else:
            print('输入错误，请重新输入：\n')
