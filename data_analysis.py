# Python Version 3.8
# @Author huanghaiyang
# Create by 2023-02-06
import os, sys, time
import sqlite3

import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl import load_workbook

# 数据库文件的绝对路径，暂时未用到
# BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# db_path = os.path.join(BASE_DIR, "answer_data.db")


# 初始化建立数据库
def createDataBase():
    cn = sqlite3.connect("./answer_data.db")

    cn.execute('''CREATE TABLE IF NOT EXISTS TB_answer
         (ID INTEGER PRIMARY KEY AUTOINCREMENT,
         batchno  varchar(32),  --导入的答卷名称
         answer_seqno varchar(32),  --答题序号
         custom_id varchar(32),     --用户ID
         answer_time varchar(32),   --提交答卷时间
         answer_time_consume INTEGER, --所用时间 单位秒
         score INTEGER,               --总分
         custom_name varchar(32),     --您的姓名：
         party_branch varchar(32),    --所在支部：1-5为第一到第五党支部，6为安金所党支部
         custom_type varchar(32)      --人员类别：1党员，2预备党员，3发展对象，4入党积极丰子，5入党申请人，6群众
         );''')

    cn.commit()
    cn.close()
    print("创建数据库answer_data.db以及TB_answer表完成！")


# 清空数据
def delteData():
    cn = sqlite3.connect("./answer_data.db")
    cn.execute("delete from TB_answer")
    cn.commit()
    cn.close()
    print("清空数据完成！\n")


# 数据库查询的结果导出为excel文件
def exportExcel(sql, title, filename):
    """
    :param sql: 待执行的查询语句
    :param title: 表头
    :param filename: 保存文件名
    :return:
    """
    print("导出数据中......")
    time_stamp = time.time()
    wb = openpyxl.Workbook()  # 新建excel文件
    ws = wb.worksheets[0]
    for i in range(len(title)):
        ws.cell(1, i + 1).value = title[i]  # 将表头写入excel文件第一行

    # 数据查询结果写入数据
    # 温馨提示：openpyxl中首行和首列从1开始计算
    conn = sqlite3.connect("./answer_data.db")
    c = conn.cursor()
    cursor = c.execute(sql)
    rows = cursor.fetchall()
    for i in range(len(rows)):
        row = rows[i]
        for j in range(len(row)):
            ws.cell(i + 2, j + 1).value = row[j]
    # 保存excel文件
    excel_filename = "./export/export_" + filename + str(time_stamp) + ".xlsx"
    wb.save(excel_filename)
    conn.close()
    print("导出数据中完成！excel文件：%s" % excel_filename)


# 解析excel文件并将其存储到sqlite
def importExcel():
    # 获取目录中文件名，并判断是否为excel文件
    file_names = os.listdir("./import")
    for i in range(len(file_names)):
        file_type = file_names[i].split(".")[-1]
        if file_type not in ["xlsx", "xls"]:
            del file_names[i]  # file_names中删除非excel格式文件

    print("\nimport文件夹发现如下excel文件：")
    for filename in file_names:
        print("excel文件：%s" % filename)
    is_import = input("\n是否要导入上述excel文件? Y(y) / N(n):")
    if is_import == 'Y' or is_import == 'y':
        print("导入数据中......")
        for filename in file_names:  # 遍历import文件，逐个导入数据库
            # filename = "import/202301.xlsx"
            wb = load_workbook("import/" + filename)
            ws = wb.worksheets[0]
            rows = ws.max_row  # excel行数
            # column = ws.max_column  # excel列数
            cn = sqlite3.connect("./answer_data.db")

            for i in range(2, rows + 1):
                if ws.cell(i, 1).value is not None:
                    sql = "insert into TB_answer (batchno,answer_seqno,custom_id,answer_time," \
                          "answer_time_consume,score,custom_name,party_branch,custom_type)" \
                          " values ('%s','%s','%s','%s','%s','%s','%s','%s','%s')" \
                          % (filename.split(".")[0], ws.cell(i, 1).value, ws.cell(i, 2).value,
                             ws.cell(i, 3).value, ws.cell(i, 4).value.replace('秒', ''), ws.cell(i, 8).value,
                             ws.cell(i, 9).value, ws.cell(i, 10).value, ws.cell(i, 11).value)
                    # print(sql)
                    cn.execute(sql)
            cn.commit()
            cn.close()
            print("excel文件：", filename, "导入完成！\n")
        print("导入数据成功！")


# 查询原始答题数据
def query_all():
    batchno = input("请输入答题批次号 (不输入查全部)：")
    party_branch = input("请输入党支部编号, \n"
                         "1-5为第一到第五党支部，6为安金所党支部 (不输入查全部)：")

    if batchno == '' and party_branch == '':
        sql = "SELECT batchno, answer_seqno, custom_id, answer_time, " \
              "answer_time_consume,score,custom_name,party_branch,custom_type from TB_answer"
    elif batchno and party_branch == '':
        sql = "SELECT batchno, answer_seqno, custom_id, answer_time, " \
              "answer_time_consume,score,custom_name,party_branch,custom_type from TB_answer " \
              "where batchno='%s'" % batchno
    elif batchno == '' and party_branch:
        sql = "SELECT batchno, answer_seqno, custom_id, answer_time, " \
              "answer_time_consume,score,custom_name,party_branch,custom_type from TB_answer " \
              "where party_branch='%s'" % party_branch
    else:
        sql = "SELECT batchno, answer_seqno, custom_id, answer_time, " \
              "answer_time_consume,score,custom_name,party_branch,custom_type from TB_answer " \
              "where batchno='%s' and party_branch='%s'" % (batchno, party_branch)
    # print(sql)

    conn = sqlite3.connect("./answer_data.db")
    c = conn.cursor()
    print("\n========查询原始答题数据========")
    print("| 答卷名称 | 答题序号 | 用户ID | 提交答卷时间 | 所用时间 单位秒 | 总分 | 您的姓名 | 所在支部 | 人员类别 |")
    cursor = c.execute(sql)
    for row in cursor:
        print(row)
    print("\n====END====")
    conn.close()

    # 数据导出为excel文件
    title = ['答卷名称', '答题序号', '用户ID', '提交答卷时间', '所用时间(单位秒)', '总分', '您的姓名', '所在支部', '人员类别']
    is_export = input('是否导出该数据为excel格式？ Y(y) / N(n):')
    if is_export == 'Y' or is_export == 'y':
        exportExcel(sql, title, '原始答题数据')
    elif is_export == 'N' or is_export == 'n':
        pass
    else:
        print('输入错误，请重新输入：\n')


# 查询答题排名统计
def query_answer_rank():
    conn = sqlite3.connect("./answer_data.db")
    c = conn.cursor()
    # 排名sql
    sql = "SELECT row_number() over(PARTITION BY batchno order BY batchno,answer_time_consume) `NO`," \
          "batchno,custom_id,score,min(a1.answer_time_consume) AS answer_time_consume " \
          "FROM TB_answer a1 " \
          "WHERE a1.score=" \
          "(SELECT max(score) FROM TB_answer t  GROUP BY t.custom_id) " \
          "GROUP BY batchno,custom_id " \
          "ORDER BY batchno,score desc,answer_time_consume asc"
    print("\n========查询答题排名统计========")
    print("| 排名 | 答卷批次 | 姓名 | 分数 | 耗时 |")

    cursor = c.execute(sql)
    for row in cursor:
        print(row)
    print("\n====END====")
    conn.close()

    # 数据导出为excel文件
    title = ['排名', '答卷批次', '姓名', '分数', '耗时']
    is_export = input('是否导出该数据为excel格式？ Y(y) / N(n):')
    if is_export == 'Y' or is_export == 'y':
        exportExcel(sql, title, '答题排名统计')
    elif is_export == 'N' or is_export == 'n':
        pass
    else:
        print('输入错误，请重新输入：\n')


def query_basic_info():
    conn = sqlite3.connect("./answer_data.db")
    c = conn.cursor()
    cursor = c.execute("SELECT row_number() over(order BY t.batchno) `no`, batchno FROM TB_answer t GROUP BY batchno;")
    print("\n========答题次数统计========")
    print("| 序号 | 答卷批次 |")
    for row in cursor:
        print(row)

    sql = "SELECT batchno, case party_branch " \
          "WHEN '1' THEN '第一党支部' " \
          "WHEN '2' THEN '第二党支部' " \
          "WHEN '3' THEN '第三党支部'" \
          "WHEN '4' THEN '第四党支部'" \
          "WHEN '5' THEN '第五党支部'" \
          "WHEN '6' THEN '安金所党支部' end," \
          "count(DISTINCT(custom_id)) AS num from TB_answer GROUP BY batchno,party_branch"
    cursor = c.execute(sql)
    print("\n========答题党支部情况数据统计========")
    print("| 答卷名称 | 党支部 | 参加人数 |")
    for row in cursor:
        print(row)
    print("\n====END====")
    conn.close()

    # 数据导出为excel文件
    title = ['答卷名称', '党支部', '参加人数']
    is_export = input('是否导出该数据为excel格式？ Y(y) / N(n):')
    if is_export == 'Y' or is_export == 'y':
        exportExcel(sql, title, '答题党支部情况数据统计')
    elif is_export == 'N' or is_export == 'n':
        pass
    else:
        print('输入错误，请重新输入：\n')


if __name__ == "__main__":
    while True:
        print('\n                     欢迎使用数据统计程序\n')
        print('===================================================================')
        print('*   1.导入数据                                                       ')
        print('*   2.查询原始答题数据                                                ')
        print('*   3.查询总体统计信息                                                ')
        print('*   4.查询答题排名统计                                                ')
        print('*   99.清空数据                                                     ')
        print('*   0.退出                                                          ')
        print('*                                                                  ')
        print('*                                               Version:1.0        ')
        print('*                                          Create by Huanghaiyang  ')
        print('===================================================================')
        print('\n')
        print('请输入命令编号')

        cmd_no = input('Please input command No:')

        if cmd_no == '1':
            importExcel()
            os.system("pause")
            # sys.exit()
        if cmd_no == '2':
            query_all()
            os.system("pause")
            # sys.exit()
        if cmd_no == '3':
            query_basic_info()
            os.system("pause")
            # sys.exit()
        if cmd_no == '4':
            query_answer_rank()
            os.system("pause")
            # sys.exit()
        if cmd_no == '99':
            delteData()
            os.system("pause")
        elif cmd_no == '0':
            break
        else:
            print('输入错误，请重新输入：\n')
